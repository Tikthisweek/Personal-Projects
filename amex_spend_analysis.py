#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Mar 12 14:00:14 2024

@author: KartikPatel
"""

import pandas as pd
import matplotlib.pyplot as plt
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, NamedStyle

monthly_statement= pd.read_excel("/Users/KartikPatel/Downloads/Activity.xlsx", header=6)


def zero_out_payments(statement):
    statement.loc[statement["Amount"] < 0, "Amount"] = 0
    return statement


def cleanup(statement):
    for name in statement.Description:
        word = name.split(" ")
        statement.loc[statement["Description"] == name, "Description"] = word[0]
        word = []
    return statement

monthly_statement = zero_out_payments(monthly_statement)
monthly_statement = cleanup(monthly_statement)


def spendperday(statement):
    spend_daily = pd.DataFrame()
    spend_daily = statement.groupby("Date")["Amount"].sum().reset_index()
    spend_daily.columns = ["Date", "Total_Spent"]
    return spend_daily

def spendbyvendor(statement):
    vendor_spend = pd.DataFrame()
    vendor_spend= statement.groupby("Description")["Amount"].sum().reset_index()
    vendor_spend.columns = ["Vendor", "Amount"]
    total_spend = vendor_spend.Amount.sum()
    new_row = pd.DataFrame({"Vendor": "Total_Spend", "Amount": [total_spend]})
    vendor_spend = pd.concat([vendor_spend, new_row], ignore_index=True)
    return vendor_spend

def format_numeric(x):
    if isinstance(x, (int, float)):
        return "{:.2f}".format(x)
    return x


vendorspendlist = spendbyvendor(monthly_statement)
dailyspendlist = spendperday(monthly_statement)

vendorspendlist.sort_values(by="Amount")

Excel_filepath = (
    "/Users/KartikPatel/Desktop/Automate_The_Boring_Stuff/monthly_statement/Cycle_Summary.xlsx"
)

statement_month = monthly_statement.Date[0].split("/")[0]
statement_year = monthly_statement.Date[0].split("/")[2]

with pd.ExcelWriter(Excel_filepath, mode="a") as writer:
    name = statement_month + "-" + statement_year + " Vendors"
    vendorspendlist .to_excel(writer, sheet_name=name, index=False)

stats = monthly_statement.describe()
stats = stats.applymap(format_numeric)

with pd.ExcelWriter(Excel_filepath, mode="a") as writer:
    statname = statement_month + "-" + statement_year + " Stats"
    stats.to_excel(writer, sheet_name=statname, index=True)

with pd.ExcelWriter(Excel_filepath, mode="a") as writer:
    name = statement_month + "-" + statement_year + " Daily Spend"
    dailyspendlist.to_excel(writer, sheet_name=name, index=False)

workbook = openpyxl.load_workbook(Excel_filepath)
border_style = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


vendorsheet = workbook[name]
for cell in vendorsheet[1]:
    cell.font = Font(bold=True)


for row in vendorsheet.iter_rows(min_col=1, max_col=2):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style

for row in vendorsheet.iter_rows(min_row=2, min_col=2):
    for cell in row:
        cell.number_format = '"$"#,##0.00'


statssheet = workbook[statname]
statssheet.cell(row=1, column=1, value="Stat")
for cell in statssheet[1]:
    cell.font = Font(bold=True)

for row in statssheet.iter_rows(min_row=2, max_row=9):
    for cell in row:
        cell.font = Font(bold=False)

for row in statssheet.iter_rows(min_col=1, max_col=2):
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_style

for row in statssheet.iter_rows(min_row=2, min_col=2):
    for cell in row:
        cell.number_format = '"$"#,##0.00'

workbook.save(Excel_filepath)

plt.bar(
    vendorspendlist.loc[vendorspendlist["Amount"] > 100, "Vendor"],
    vendorspendlist.loc[vendorspendlist["Amount"] > 100, "Amount"],
)
plt.xlabel("Vendor", fontsize=8)  # Adjust the font size here
plt.ylabel("Amount ($)")
plt.title("Amount Spent by Vendor (Amount > $100)")
plt.xticks(rotation=45, ha="right", fontsize=8)  # Adjust the font size here
plt.tight_layout()  # Adjust layout to prevent labels from overlapping

min_date = str(monthly_statement.Date.min())
min_date = min_date.replace("/", "_")
max_date = str(monthly_statement.Date.max())
max_date = max_date.replace("/", "_")


folder_path = (
    "/Users/KartikPatel/Desktop/Automate_The_Boring_Stuff/monthly_statement/monthly_statement_spend_graphs"
)
file_name = min_date + " to " + max_date + ".png"
file_path = folder_path + "/" + file_name
plt.savefig(file_path)


os.system(f"open {Excel_filepath}")
remove_file = "/Users/KartikPatel/Downloads/Activity.xlsx"
os.remove(remove_file)
